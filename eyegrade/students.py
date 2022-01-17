# Eyegrade: grading multiple choice questions with a webcam
# Copyright (C) 2010-2021 Jesus Arias Fisteus
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful, but
# WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the GNU
# General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see
# <https://www.gnu.org/licenses/>.
#

import re
import csv
import itertools
import enum

import openpyxl

from . import utils

re_email = r"^[a-zA-Z0-9._%-\+]+@[a-zA-Z0-9._%-]+.[a-zA-Z]{2,6}$"
_re_email = re.compile(re_email)
_re_student_id = re.compile(r"^[0-9]+$")


class Student:
    def __init__(
        self,
        student_id,
        full_name,
        first_name,
        last_name,
        email,
        db_id=None,
        group_id=None,
        sequence_num=None,
        is_in_database=False,
    ):
        if full_name and (first_name or last_name):
            raise ValueError("Full name incompatible with first / last name")
        self.db_id = db_id
        self.student_id = student_id
        self.full_name = full_name
        self.first_name = first_name
        self.last_name = last_name
        self.email = email
        self.group_id = group_id
        self.sequence_num = sequence_num
        self.is_in_database = is_in_database
        self.is_duplicate = False

    @property
    def name(self):
        if self.full_name:
            return self.full_name
        elif self.last_name:
            if self.first_name:
                return "{0} {1}".format(self.first_name, self.last_name)
            else:
                return self.last_name
        elif self.first_name:
            return self.first_name
        else:
            return ""

    @property
    def last_comma_first_name(self):
        if self.last_name:
            if self.first_name:
                return "{0}, {1}".format(self.last_name, self.first_name)
            else:
                return self.last_name
        else:
            return self.name

    @property
    def id_and_name(self):
        if self.name:
            return " ".join((self.student_id, self.name))
        else:
            return self.student_id

    @property
    def name_or_id(self):
        if self.name:
            return self.name
        elif self.student_id:
            return self.student_id
        else:
            return ""

    def __lt__(self, other):
        return self.student_id < other.student_id

    def __eq__(self, other):
        if isinstance(other, Student):
            return self.student_id == other.student_id
        else:
            return False

    def __str__(self):
        return "student: " + self.id_and_name


class StudentGroup:
    def __init__(self, identifier, name):
        self.identifier = identifier
        self.name = name

    def __str__(self):
        return "Group #{0.identifier} ({0.name})".format(self)


class GroupListing:
    def __init__(self, group, students):
        self.group = group
        self.students = list(students)
        self.parent = None
        self._students_dict = {s.student_id: s for s in students}

    def student(self, student_id):
        return self._students_dict.get(student_id, None)

    def add_students(self, student_list):
        if student_list:
            if self.parent is not None:
                duplicates = self.parent.find_duplicates(student_list)
            else:
                duplicates = self.find_duplicates(student_list)
            if not duplicates:
                if student_list[0].sequence_num is None:
                    self._update_sequence_num(student_list)
                self.students.extend(student_list)
                for student in student_list:
                    student.group_id = self.group.identifier
                self._students_dict.update({s.student_id: s for s in student_list})
            else:
                raise DuplicateStudentIdException(duplicates)

    def remove_students(self, students):
        for student in students:
            if student.student_id in self._students_dict:
                del self._students_dict[student.student_id]
                self.students.remove(student)

    def rename(self, new_name):
        self.group.name = new_name

    def find_duplicates(self, students):
        non_duplicates, duplicates = _duplicate_student_ids(students)
        duplicates.extend([s for s in non_duplicates if s.student_id in self])
        return duplicates

    def __len__(self):
        return len(self.students)

    def __getitem__(self, key):
        return self.students[key]

    def __iter__(self):
        return iter(self.students)

    def __contains__(self, student_id):
        return student_id in self._students_dict

    def __str__(self):
        return "GroupListing({}, {} students)".format(self.group, len(self.students))

    def _update_sequence_num(self, students):
        if len(self.students) > 0:
            first_num = 1 + max(s.sequence_num for s in self.students)
        else:
            first_num = 1
        for i, student in enumerate(students):
            student.sequence_num = first_num + i


class StudentListings:
    def __init__(self):
        self.listings = []
        self.max_group_id = -1
        self._sorted_students = None

    def add_listing(self, listing):
        duplicates = self.find_duplicates(listing.students)
        if not duplicates:
            listing.parent = self
            self.listings.append(listing)
            group_id = listing.group.identifier
            if group_id > self.max_group_id:
                self.max_group_id = group_id
        else:
            raise DuplicateStudentIdException(duplicates)

    def create_listing(self, group):
        if group.identifier is None:
            group.identifier = self.max_group_id + 1
        listing = GroupListing(group, [])
        self.add_listing(listing)
        return listing

    def remove_at(self, index):
        del self.listings[index]

    def iter_students(self):
        return itertools.chain(*self.listings)

    def sorted_students(self, key=lambda x: x.student_id):
        return sorted([student for student in self.iter_students()], key=key)

    def student(self, student_id):
        student = None
        for listing in self.listings:
            student = listing.student(student_id)
            if student is not None:
                break
        return student

    def listing_by_group_id(self, group_id):
        for listing in self.listings:
            if listing.group.identifier == group_id:
                return listing
        raise KeyError(group_id)

    def find_duplicates(self, students):
        non_duplicates, duplicates = _duplicate_student_ids(students)
        duplicates.extend([s for s in non_duplicates if s.student_id in self])
        return duplicates

    def __len__(self):
        return len(self.listings)

    def __getitem__(self, key):
        return self.listings[key]

    def __contains__(self, student_id):
        for listing in self.listings:
            if student_id in listing:
                return True
        return False

    def __str__(self):
        return "Studentlistings({} groups)".format(len(self.listings))


def _duplicate_student_ids(students):
    non_duplicates = []
    duplicates = []
    student_ids_set = set()
    for student in students:
        if student.student_id not in student_ids_set:
            student_ids_set.add(student.student_id)
            non_duplicates.append(student)
        else:
            duplicates.append(student)
    return non_duplicates, duplicates


class CantRemoveGroupException(utils.EyegradeException):
    def __init__(self, message):
        super().__init__(message)


class DuplicateStudentIdException(utils.EyegradeException):
    def __init__(self, duplicates):
        super().__init__("Some ids are already in the student listings")
        self.duplicates = duplicates


class StudentReader:
    def __init__(self, file_name, column_map=None):
        self.file_name = file_name
        self.column_map = column_map
        # To be overwritten by subclasses:
        self.iterator = iter([])

    @staticmethod
    def create(file_name):
        if file_name.endswith(".xlsx"):
            return XLSXStudentReader(file_name)
        else:
            return CSVStudentReader(file_name)

    def students(self):
        first_line = True
        for row in self.iterator:
            if not StudentReader._row_is_empty(row):
                if self.column_map is None:
                    self.column_map = StudentColumnMap.guess_map(row)
                    if not self.column_map.is_valid():
                        if first_line:
                            first_line = False
                            self.column_map = None
                            continue
                        else:
                            raise utils.EyegradeException("", key="error_student_list")
                try:
                    student = self.column_map.student(row)
                    yield student
                except utils.EyegradeException:
                    if not first_line:
                        raise
                first_line = False

    @staticmethod
    def _row_is_empty(row):
        for element in row:
            if element is not None and element != "":
                return False
        return True


class CSVStudentReader(StudentReader):
    def __init__(self, file_name):
        super().__init__(file_name)
        self.file = None

    def __enter__(self):
        self.file = open(self.file_name, newline="")
        try:
            dialect = csv.Sniffer().sniff(self.file.read(1024))
        except csv.Error:
            dialect = csv.excel_tab
        self.file.seek(0)
        self.iterator = csv.reader(self.file, dialect=dialect)
        return self

    def __exit__(self, exception_type, exception_value, traceback):
        self.file.close()


class XLSXStudentReader(StudentReader):
    def __init__(self, file_name):
        super().__init__(file_name)
        self.workbook = None
        self.iterator = None

    def __enter__(self):
        self.workbook = openpyxl.load_workbook(self.file_name, read_only=True)
        self.iterator = self.iter_rows(self.workbook.active)
        return self

    def __exit__(self, exception_type, exception_value, traceback):
        self.workbook.close()

    def iter_rows(self, work_sheet):
        for row in work_sheet.iter_rows():
            yield tuple(cell.value for cell in row)


def read_students(file_name):
    """Reads the list of students from a file.

    Formats allowed: CSV-formatted file (tab-separated) and Excel 2010 (.xslx)

    Returns the results as a list of Student objects.

    """
    with StudentReader.create(file_name) as reader:
        return list(reader.students())


class StudentColumn(enum.Enum):
    ID = 1
    FULL_NAME = 2
    FIRST_NAME = 3
    LAST_NAME = 4
    NAME = 5
    EMAIL = 6
    SEQUENCE_NUM = 7
    UNKNOWN = 8


ATTR_NAME = {
    StudentColumn.ID: "student_id",
    StudentColumn.FULL_NAME: "full_name",
    StudentColumn.FIRST_NAME: "first_name",
    StudentColumn.LAST_NAME: "last_name",
    StudentColumn.NAME: "name",
    StudentColumn.EMAIL: "email",
    StudentColumn.SEQUENCE_NUM: "sequence_num",
    StudentColumn.UNKNOWN: "-",
}


class StudentColumnMap:
    def __init__(self, num_columns=None, columns=None):
        if not ((num_columns is None) ^ (columns is None)):
            raise ValueError("num_columns or columns required, but not both")
        if num_columns is not None:
            self.columns = [StudentColumn.UNKNOWN] * num_columns
        else:
            self.columns = list(columns)

    def set_column(self, index, column):
        if column not in self.columns:
            self.columns[index] = column
            return True
        else:
            return False

    def resolve(self):
        # Identify first name / last name / full name columns
        # They are marked as unkown until now
        num_columns = len(self.columns)
        for i in range(num_columns):
            if self.columns[i] == StudentColumn.UNKNOWN:
                if i == num_columns - 1 or self.columns[i + 1] != StudentColumn.UNKNOWN:
                    self.columns[i] = StudentColumn.FULL_NAME
                else:
                    self.columns[i] = StudentColumn.FIRST_NAME
                    self.columns[i + 1] = StudentColumn.LAST_NAME
                break
        for i, col in reversed(list(enumerate(self.columns))):
            if col != StudentColumn.UNKNOWN:
                break
        self.columns = self.columns[: i + 1]

    def is_valid(self):
        return StudentColumn.ID in self.columns

    def student(self, row):
        num_columns = len(self.columns)
        if len(row) < num_columns:
            raise utils.EyegradeException(
                "Row with not enough columns", key="error_student_list"
            )
        student = Student("", "", "", "", "")
        for i, item in enumerate(row[:num_columns]):
            if self.columns[i] != StudentColumn.UNKNOWN:
                value = str(item)
                self._check_value(self.columns[i], value)
                attr_name = ATTR_NAME[self.columns[i]]
                setattr(student, attr_name, value)
        return student

    def data(self, index, student):
        column = self.columns[index]
        if column != StudentColumn.UNKNOWN:
            attr_name = ATTR_NAME[column]
            return getattr(student, attr_name)
        else:
            return ""

    def normalize(self):
        normal_order = [
            StudentColumn.ID,
            StudentColumn.FIRST_NAME,
            StudentColumn.LAST_NAME,
            StudentColumn.FULL_NAME,
            StudentColumn.EMAIL,
        ]
        reordered_columns = [
            column for column in normal_order if column in self.columns
        ]
        return StudentColumnMap(columns=reordered_columns)

    def to_full_name(self):
        # It will raise ValueError if first or last name aren't present
        index_first = self.columns.index(StudentColumn.FIRST_NAME)
        index_last = self.columns.index(StudentColumn.LAST_NAME)
        new_columns = list(self.columns)
        if index_first < index_last:
            new_columns[index_first] = StudentColumn.FULL_NAME
            del new_columns[index_last]
        else:
            new_columns[index_last] = StudentColumn.FULL_NAME
            del new_columns[index_first]
        return StudentColumnMap(columns=new_columns)

    def __str__(self):
        return (
            "StudentColumnMap <"
            + ", ".join(ATTR_NAME[column] for column in self.columns)
            + ">"
        )

    def __len__(self):
        return len(self.columns)

    def __getitem__(self, index):
        return self.columns[index]

    def __contains__(self, column):
        return column in self.columns

    @staticmethod
    def guess_map(row):
        column_map = StudentColumnMap(num_columns=len(row))
        for i, item in enumerate(row):
            value = str(item)
            if _re_student_id.match(value):
                column = StudentColumn.ID
            elif _re_email.match(value):
                column = StudentColumn.EMAIL
            else:
                column = StudentColumn.UNKNOWN
            column_map.set_column(i, column)
        column_map.resolve()
        return column_map

    def _check_value(self, column, value):
        if column == StudentColumn.ID:
            if not _re_student_id.match(value):
                raise utils.EyegradeException(
                    "Wrong id in student list: " + value, key="error_student_list"
                )
        elif column == StudentColumn.EMAIL:
            if not _re_email.match(value):
                raise utils.EyegradeException(
                    "Wrong email in student list: " + value, key="error_student_list"
                )


utils.EyegradeException.register_error(
    "error_student_list", "The syntax of the student list isn't correct."
)
